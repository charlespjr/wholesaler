#!/usr/bin/env python3
"""
build_deal_packages.py
======================
Extract publicly accessible property photos from Zillow / Redfin listing pages
and assemble a professional wholesale deal-package PDF per property.

Usage
-----
Batch (CSV or Excel):
    python build_deal_packages.py --input properties.csv --output output --max-photos 15

Single listing:
    python build_deal_packages.py --url "LISTING_URL" --address "123 Main St, City, ST 00000" --output output

This tool only reads images the public listing page already serves. It does not
log in, solve CAPTCHAs, or defeat bot protection. If a page blocks an anonymous
browser, the PDF is still produced from the available property data with a clear
"Listing photos could not be retrieved." note.
"""
from __future__ import annotations

import argparse
import csv
import logging
import os
import sys
from typing import Dict, List, Optional

from extractors import detect_source, extract_photo_urls
from images import collect_photos
from models import Property
from pdfgen import Brand, build_pdf

log = logging.getLogger("deal_packages")

# Default company branding (Paragon letterhead). Logo auto-located relative to
# this file at ../assets/paragon_logo.png unless --logo is given.
PARAGON_BRAND = Brand(
    company="Paragon Government Solutions LLC",
    line2="11166 Fairfax Blvd, Suite 500, Fairfax, VA 22030  |  (888) 495-6935  |  charlesp@paragongovsolutions.net",
    line3="UEI: FSCZBK8CBV82  |  CAGE Code: 9WX69  |  www.paragongovsolutions.net",
)
_DEFAULT_LOGO = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "assets", "paragon_logo.png")


def resolve_brand(no_brand: bool, logo_path: str = "") -> Optional[Brand]:
    if no_brand:
        return None
    brand = Brand(company=PARAGON_BRAND.company, line2=PARAGON_BRAND.line2, line3=PARAGON_BRAND.line3)
    candidate = logo_path or _DEFAULT_LOGO
    brand.logo_path = candidate if os.path.exists(candidate) else None
    if not brand.logo_path:
        log.warning("Brand logo not found at %s - building branded text footer without logo.", candidate)
    return brand


def _setup_logging(verbose: bool) -> None:
    logging.basicConfig(
        level=logging.DEBUG if verbose else logging.INFO,
        format="%(asctime)s  %(levelname)-7s %(name)s: %(message)s",
        datefmt="%H:%M:%S",
    )


# -----------------------------------------------------------------------------
# Input reading
# -----------------------------------------------------------------------------
def read_rows(path: str) -> List[Dict]:
    ext = os.path.splitext(path)[1].lower()
    if ext in (".xlsx", ".xls"):
        return _read_excel(path)
    return _read_csv(path)


def _normalize_headers(fieldnames) -> Dict[str, str]:
    """Map loose header spellings to our canonical column names."""
    out = {}
    for f in fieldnames or []:
        key = f.strip().lower().replace(" ", "_")
        out[f] = key
    return out


def _read_csv(path: str) -> List[Dict]:
    with open(path, newline="", encoding="utf-8-sig") as fh:
        reader = csv.DictReader(fh)
        hmap = _normalize_headers(reader.fieldnames)
        rows = []
        for raw in reader:
            rows.append({hmap.get(k, k): v for k, v in raw.items()})
        return rows


def _read_excel(path: str) -> List[Dict]:
    try:
        import pandas as pd
        df = pd.read_excel(path, dtype=str).fillna("")
        df.columns = [str(c).strip().lower().replace(" ", "_") for c in df.columns]
        return df.to_dict(orient="records")
    except ImportError:
        # Fall back to openpyxl directly if pandas is unavailable.
        from openpyxl import load_workbook
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        headers = [str(h).strip().lower().replace(" ", "_") if h is not None else ""
                   for h in next(rows_iter)]
        out = []
        for r in rows_iter:
            out.append({headers[i]: ("" if v is None else str(v)) for i, v in enumerate(r)
                        if i < len(headers)})
        return out


# -----------------------------------------------------------------------------
# Per-property processing
# -----------------------------------------------------------------------------
def process_property(prop: Property, output_dir: str, max_photos: int,
                     headless: bool, min_width: int, brand: Optional[Brand] = None) -> Dict:
    report = {
        "property_address": prop.property_address,
        "listing_url": prop.listing_url,
        "source": detect_source(prop.listing_url) if prop.listing_url else "",
        "photos_found": 0,
        "photos_downloaded": 0,
        "pdf_created": False,
        "error_message": "",
    }
    photos = []
    photos_note = ""

    try:
        if prop.listing_url:
            urls, source = extract_photo_urls(prop.listing_url, headless=headless,
                                              max_photos=max_photos)
            prop.source = source
            report["source"] = source
            report["photos_found"] = len(urls)
            if urls:
                photo_dir = os.path.join(output_dir, "photos", prop.photo_folder_name())
                photos = collect_photos(urls, photo_dir, referer=prop.listing_url,
                                        max_photos=max_photos, min_width=min_width)
                report["photos_downloaded"] = len(photos)
            if not photos:
                photos_note = "Listing photos could not be retrieved."
        else:
            photos_note = "No listing URL provided; package built from property data only."
    except Exception as exc:  # noqa: BLE001 - keep batch going
        log.exception("Photo extraction failed for %s", prop.title)
        report["error_message"] = f"photo_extraction: {exc}"
        photos_note = "Listing photos could not be retrieved."

    # PDF is always attempted, even with zero photos.
    try:
        pkg_dir = os.path.join(output_dir, "deal_packages")
        os.makedirs(pkg_dir, exist_ok=True)
        out_path = os.path.join(pkg_dir, prop.package_basename())
        build_pdf(prop, photos, out_path, photos_note=photos_note, brand=brand)
        report["pdf_created"] = True
    except Exception as exc:  # noqa: BLE001
        log.exception("PDF build failed for %s", prop.title)
        report["error_message"] = (report["error_message"] + f"; pdf: {exc}").strip("; ")

    return report


def write_report(rows: List[Dict], output_dir: str) -> str:
    path = os.path.join(output_dir, "processing_report.csv")
    cols = ["property_address", "listing_url", "source", "photos_found",
            "photos_downloaded", "pdf_created", "error_message"]
    with open(path, "w", newline="", encoding="utf-8") as fh:
        w = csv.DictWriter(fh, fieldnames=cols)
        w.writeheader()
        for r in rows:
            w.writerow({c: r.get(c, "") for c in cols})
    log.info("Processing report: %s", path)
    return path


# -----------------------------------------------------------------------------
# CLI
# -----------------------------------------------------------------------------
def main(argv=None) -> int:
    ap = argparse.ArgumentParser(description="Build wholesale deal-package PDFs with listing photos.")
    src = ap.add_argument_group("input (choose one)")
    src.add_argument("--input", help="CSV or Excel file, one property per row.")
    src.add_argument("--url", help="A single Zillow or Redfin listing URL.")
    ap.add_argument("--address", default="", help="Property address (used with --url).")
    ap.add_argument("--output", default="output", help="Output directory (default: output).")
    ap.add_argument("--max-photos", type=int, default=15, help="Max photos per property (default: 15).")
    ap.add_argument("--min-width", type=int, default=600, help="Skip images narrower than this (default: 600).")
    ap.add_argument("--no-headless", action="store_true", help="Show the browser window (debugging).")
    ap.add_argument("--no-brand", action="store_true", help="Disable Paragon letterhead branding.")
    ap.add_argument("--logo", default="", help="Path to a logo PNG (overrides the default Paragon logo).")
    ap.add_argument("--verbose", "-v", action="store_true", help="Verbose logging.")
    args = ap.parse_args(argv)

    _setup_logging(args.verbose)
    if not args.input and not args.url:
        ap.error("provide --input <file> or --url <listing_url>")

    os.makedirs(args.output, exist_ok=True)
    headless = not args.no_headless

    properties: List[Property] = []
    if args.url:
        properties.append(Property.from_row({
            "property_address": args.address, "listing_url": args.url,
        }))
    else:
        try:
            rows = read_rows(args.input)
        except FileNotFoundError:
            log.error("Input file not found: %s", args.input)
            return 2
        for raw in rows:
            # ignore fully-empty rows
            if not any(str(v).strip() for v in raw.values()):
                continue
            properties.append(Property.from_row(raw))

    if not properties:
        log.error("No properties to process.")
        return 2

    brand = resolve_brand(args.no_brand, args.logo)
    log.info("Processing %d propert%s …  (branding: %s)", len(properties),
             "y" if len(properties) == 1 else "ies",
             "Paragon" if brand else "off")
    reports = []
    for i, prop in enumerate(properties, 1):
        log.info("[%d/%d] %s", i, len(properties), prop.title)
        reports.append(process_property(prop, args.output, args.max_photos, headless,
                                        args.min_width, brand=brand))

    write_report(reports, args.output)
    ok = sum(1 for r in reports if r["pdf_created"])
    with_photos = sum(1 for r in reports if r["photos_downloaded"] > 0)
    log.info("Done. PDFs created: %d/%d  (with photos: %d).  Output: %s",
             ok, len(properties), with_photos, os.path.abspath(args.output))
    return 0


if __name__ == "__main__":
    sys.exit(main())
